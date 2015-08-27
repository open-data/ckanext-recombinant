import openpyxl

from ckanext.recombinant.plugins import get_table
from ckanext.recombinant.errors import RecombinantException
from ckanext.recombinant.datatypes import data_store_type

boolean_validator = openpyxl.worksheet.datavalidation.DataValidation(
    type="list", formula1='"FALSE,TRUE"', allow_blank=True)

def xls_template(dataset_type, org):
    """
    return an openpyxl.Workbook object containing the sheet and header fields
    for passed dataset_type and org.
    """
    t = get_table(dataset_type)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = t['xls_sheet_name']

    ref = book.create_sheet(title='reference')
    ref.append([u'field', u'key', u'value'])

    sheet.add_data_validation(boolean_validator)

    for n, key in enumerate(t['xls_organization_info']):
        c = sheet.cell(row=1, column=n + 1)
        for e in org['extras']:
            if e['key'] == key:
                c.value = e['value']
                break
        else:
            c.value = org.get(key, '')
        apply_styles(t['excel_organization_style'], c)
    apply_styles(t['excel_organization_style'], sheet.row_dimensions[1])

    for n, field in enumerate(t['fields']):
        c = sheet.cell(row=2, column=n + 1)
        c.value = field['label']
        apply_styles(t['excel_header_style'], c)
        # jumping through openpyxl hoops:
        col_letter = openpyxl.cell.get_column_letter(n + 1)
        col = sheet.column_dimensions[col_letter]
        col.width = field['xls_column_width']
        # FIXME: format only below header
        col.number_format = data_store_type[field['datastore_type']].xl_format
        validation_range = '{0}3:{0}1003'.format(col_letter)
        if field['datastore_type'] == 'boolean':
            boolean_validator.ranges.append(validation_range)
        elif 'choices' in field:
            v = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1='"' + ','.join(sorted(field['choices'])) + '"',
                allow_blank=True)
            v.errorTitle = u'Invalid choice'
            v.error = u'Please enter one of the following codes:\n' + u', '.join(
                sorted(field['choices'])
                ) + '\n\nSheet "reference" shows code values.'
            sheet.add_data_validation(v)
            v.ranges.append(validation_range)

            ref.append([field['label']])
            for key, value in sorted(field['choices'].iteritems()):
                ref.append([None, key, value])
            ref.append([])
    apply_styles(t['excel_header_style'], sheet.row_dimensions[2])

    sheet.freeze_panes = sheet['A3']
    return book


def apply_styles(config, target):
    """
    apply styles from config to target

    currently supports PatternFill and Font
    """
    pattern_fill = config.get('PatternFill')
    if pattern_fill:
        target.fill = openpyxl.styles.PatternFill(**pattern_fill)
    font = config.get('Font')
    if font:
        target.font = openpyxl.styles.Font(**font)

