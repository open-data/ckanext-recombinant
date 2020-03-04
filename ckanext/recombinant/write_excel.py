# coding: utf-8
"""
Excel v3 template and data-dictionary generation code
"""

import textwrap
import string
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import NamedStyle

from ckanext.recombinant.tables import get_geno
from ckanext.recombinant.errors import RecombinantException
from ckanext.recombinant.datatypes import datastore_type
from ckanext.recombinant.helpers import (
    recombinant_choice_fields, recombinant_language_text)
from ckanext.recombinant.write_excel_v2 import (
    _populate_excel_sheet_v2, _populate_reference_sheet_v2)

from ckan.plugins.toolkit import _, h

from datetime import datetime
from decimal import Decimal

HEADER_ROW, HEADER_HEIGHT = 1, 27
CHEADINGS_ROW, CHEADINGS_HEIGHT = 2, 22
CHEADINGS_MIN_WIDTH = 10
LINE_HEIGHT = 15  # extra lines of text in same cell
CODE_ROW = 3
CSTATUS_ROW, CSTATUS_HEIGHT = 4, 6
EXAMPLE_ROW, DEFAULT_EXAMPLE_HEIGHT = 5, 15
EXAMPLE_MERGE = 'A5:B5'
FREEZE_PANES = 'C5'
DATA_FIRST_ROW, DEFAULT_DATA_HEIGHT = 6, 24
DEFAULT_DATA_NUM_ROWS = 2000
RSTATUS_COL, RSTATUS_COL_NUM = 'A', 1
RSTATUS_WIDTH = 1
RPAD_COL, RPAD_COL_NUM = 'B', 2
RPAD_WIDTH = 3
DATA_FIRST_COL, DATA_FIRST_COL_NUM = 'C', 3
ESTIMATE_WIDTH_MULTIPLE_1 = 1.3
ESTIMATE_WIDTH_MULTIPLE_1_CHARS = 20
ESTIMATE_WIDTH_MULTIPLE_2 = 1.0
EDGE_RANGE = 'A1:A4'

REF_HEADER1_ROW, REF_HEADER1_HEIGHT = 1, 27
REF_HEADER2_ROW, REF_HEADER2_HEIGHT = 2, 27
REF_FIRST_ROW = 4
REF_FIELD_NUM_COL, REF_FIELD_NUM_COL_NUM = 'A', 1
REF_FIELD_NUM_MERGE = 'A{row}:B{row}'
REF_FIELD_TITLE_HEIGHT = 24
REF_FIELD_TITLE_MERGE = 'C{row}:D{row}'
REF_KEY_COL, REF_KEY_COL_NUM = 'C', 3
REF_KEY_WIDTH = 18
REF_VALUE_COL, REF_VALUE_COL_NUM = 'D', 4
REF_VALUE_WIDTH = 114
REF_CHOICE_HEADING_HEIGHT = 24
REF_EDGE_RANGE = 'A1:A2'

DEFAULT_YEAR_MIN, DEFAULT_YEAR_MAX = '2018-50', '2018+50'

DEFAULT_EDGE_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FF336B87'},
    'Font': {'color': 'FFFFFF'}}
DEFAULT_HEADER_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FF90AFC5'},
    'Font': {'bold': True, 'size': 16}}
DEFAULT_CHEADING_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FF90AFC5'},
    'Alignment': {'wrapText': True},
    'Font': {'color': '000000', 'underline': 'single'}}
DEFAULT_EXAMPLE_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FFDDD9C4'},
    'Alignment': {'wrapText': True, 'vertical': 'top'}}
DEFAULT_ERROR_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FFC00000'},
    'Font': {'color': 'FFFFFF'}}
DEFAULT_REF_HEADER2_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FF90AFC5'},
    'Alignment': {'vertical': 'center'}}
REF_NUMBER_STYLE = {}
REF_TITLE_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FFFFFFFF'},
    'Font': {'underline': 'single'}}
REF_ATTR_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FFFFFFFF'},
    'Font': {'color': '666666'},
    'Alignment': {'vertical': 'top'}}
REF_VALUE_STYLE = {
    'Alignment': {'wrapText': True, 'vertical': 'top'}}
REF_PAPER_STYLE = {
    'PatternFill': {'patternType': 'solid', 'fgColor': 'FFFFFFFF'}}
TYPE_HERE_STYLE = {
    'Font': {'bold': True, 'size': 16}}


def excel_template(dataset_type, org):
    """
    return an openpyxl.Workbook object containing the sheet and header fields
    for passed dataset_type and org. Supports version 2 and 3 templates.
    """
    geno = get_geno(dataset_type)
    version = geno.get('template_version', 2)

    book = openpyxl.Workbook()
    sheet = book.active
    refs = []
    choice_ranges = []

    if version == 3:
        _build_styles(book, geno)
    for rnum, chromo in enumerate(geno['resources'], 1):
        if version == 2:
            _populate_excel_sheet_v2(sheet, chromo, org, refs)
        elif version == 3:
            choice_ranges.append(_populate_excel_sheet(
                book, sheet, geno, chromo, org, refs, rnum))
            sheet.protection.enabled = True
            sheet.protection.formatRows = False
            sheet.protection.formatColumns = False
        sheet = book.create_sheet()

    if version == 2:
        _populate_reference_sheet_v2(sheet, chromo, refs)
    elif version == 3:
        _populate_reference_sheet(sheet, geno, refs)
    sheet.title = 'reference'
    sheet.protection.enabled = True

    if version == 2:
        return book

    for i, (chromo, cranges) in enumerate(
            zip(geno['resources'], choice_ranges), 1):
        sheet = book.create_sheet()
        _populate_excel_e_sheet(sheet, chromo, cranges)
        sheet.title = 'e{i}'.format(i=i)
        sheet.protection.enabled = True
        sheet.sheet_state = 'hidden'

        sheet = book.create_sheet()
        _populate_excel_r_sheet(sheet, chromo)
        sheet.title = 'r{i}'.format(i=i)
        sheet.protection.enabled = True
        sheet.sheet_state = 'hidden'
    return book


def append_data(book, record_data, chromo):

    """
    fills rows of an openpyxl.Workbook with selected data from a datastore resource
    """
    sheet = book[chromo['resource_name']]
    current_row = DATA_FIRST_ROW
    for record in record_data:
        for col_num, field in template_cols_fields(chromo):
            item = datastore_type_format(record[field['datastore_id']], field['datastore_type'])
            sheet.cell(row=current_row, column=col_num).value = item
        current_row += 1

    return book


def datastore_type_format(value, datastore_type):

    numeric_types = ['money', 'year', 'int', 'bigint', 'numeric']
    if isinstance(value, list):
        item = u', '.join(unicode(e) for e in value)
    elif datastore_type == 'date':
        item = datetime.strptime(value, "%Y-%m-%d").date()
    elif datastore_type == 'timestamp':
        item = datetime.strptime(value, "%Y-%m-%d %H:%M:%S %Z")
    elif datastore_type in numeric_types:
        item = Decimal(value)
    else:
        item = value

    return item


def excel_data_dictionary(geno):
    """
    return an openpyxl.Workbook object containing the field reference
    from geno, one sheet per language
    """
    book = openpyxl.Workbook()
    sheet = book.active

    style1 = {
        'PatternFill': {
            'patternType': 'solid',
            'fgColor': 'FFFFF056'},
        'Font': {
            'bold': True}}
    style2 = {
        'PatternFill': {
            'patternType': 'solid',
            'fgColor': 'FFDFE2DB'}}

    from pylons import config
    from ckan.lib.i18n import handle_request, get_lang
    from ckan.common import c, request

    _build_styles(book, geno)
    for lang in config['ckan.locales_offered'].split():
        if sheet is None:
            sheet = book.create_sheet()

        sheet.title = lang.upper()
        # switch language (FIXME: this is harder than it should be)
        request.environ['CKAN_LANG'] = lang
        handle_request(request, c)
        choice_fields = dict(
            (f['datastore_id'], f['choices'])
            for chromo in geno['resources']
            for f in recombinant_choice_fields(chromo['resource_name']))

        refs = []
        for chromo in geno['resources']:
            for field in chromo['fields']:
                _append_field_ref_rows(refs, field, link=None)

                if field['datastore_id'] in choice_fields:
                    _append_field_choices_rows(
                        refs,
                        choice_fields[field['datastore_id']],
                        full_text_choices=False)

        _populate_reference_sheet(sheet, geno, refs)
        sheet = None

    return book


def estimate_width_from_length(length):
    range1 = max(length, ESTIMATE_WIDTH_MULTIPLE_1_CHARS)
    range2 = length - range1
    return (
        range1 * ESTIMATE_WIDTH_MULTIPLE_1 +
        range2 * ESTIMATE_WIDTH_MULTIPLE_2)

def estimate_width(text):
    return max(estimate_width_from_length(len(s)) for s in text.split('\n'))

def wrap_text_to_width(text, width):
    # assuming width > ESTIMATE_WIDTH_MULTIPLE_1_CHARS
    width -= ESTIMATE_WIDTH_MULTIPLE_1_CHARS * ESTIMATE_WIDTH_MULTIPLE_1
    cwidth = width // ESTIMATE_WIDTH_MULTIPLE_2 + ESTIMATE_WIDTH_MULTIPLE_1_CHARS
    return '\n'.join(
        '\n'.join(textwrap.wrap(line, cwidth))
        for line in text.split('\n'))


def _build_styles(book, geno):
    """
    Add styles to workbook
    """
    build_named_style(book, 'reco_edge', dict(
        DEFAULT_EDGE_STYLE, **geno.get('excel_edge_style', {})))
    build_named_style(book, 'reco_header', dict(
        DEFAULT_HEADER_STYLE, **geno.get('excel_header_style', {})))
    build_named_style(book, 'reco_header2', dict(
        DEFAULT_REF_HEADER2_STYLE, **geno.get('excel_header_style', {})))
    build_named_style(book, 'reco_cheading', dict(
        DEFAULT_CHEADING_STYLE, **geno.get('excel_column_heading_style', {})))
    build_named_style(book, 'reco_example', dict(
        DEFAULT_EXAMPLE_STYLE, **geno.get('excel_example_style', {})))
    build_named_style(book, 'reco_error', dict(
        DEFAULT_ERROR_STYLE, **geno.get('excel_error_style', {})))
    build_named_style(book, 'reco_ref_number', REF_NUMBER_STYLE)
    build_named_style(book, 'reco_ref_title', REF_TITLE_STYLE)
    build_named_style(book, 'reco_ref_attr', REF_ATTR_STYLE)
    build_named_style(book, 'reco_ref_value', REF_VALUE_STYLE)


def _populate_excel_sheet(book, sheet, geno, chromo, org, refs, resource_num):
    """
    Format openpyxl sheet for the resource definition chromo and org.
    (Version 3)

    refs - list of rows to add to reference sheet, modified
        in place from this function
    resource_num - 1-based index of resource

    returns cranges dict of {datastore_id: reference_key_range}
    """
    sheet.title = chromo['resource_name']

    cranges = {}
    data_num_rows = chromo.get('excel_data_num_rows', DEFAULT_DATA_NUM_ROWS)

    required_style = dict(
        dict(DEFAULT_EDGE_STYLE, **geno.get('excel_edge_style', {})),
        **geno.get('excel_required_style', {}))
    error_style = dict(
        DEFAULT_ERROR_STYLE, **geno.get('excel_error_style', {}))
    header_style = dict(
        DEFAULT_HEADER_STYLE, **geno.get('excel_header_style', {}))
    cheadings_style = dict(
        DEFAULT_CHEADING_STYLE,
        **geno.get('excel_column_heading_style', {}))
    example_style = dict(
        DEFAULT_EXAMPLE_STYLE, **geno.get('excel_example_style', {}))

    # create rows so we can set all heights
    for i in xrange(1, DATA_FIRST_ROW + data_num_rows):
        sheet.cell(row=i, column=1).value = None

    sheet.merge_cells(EXAMPLE_MERGE)
    fill_cell(sheet, EXAMPLE_ROW, 1, _('e.g.'), 'reco_example')

    fill_cell(
        sheet,
        DATA_FIRST_ROW,
        RPAD_COL_NUM,
        u'=IF(r{rnum}!{col}{row},"","▶")'.format(
            rnum=resource_num,
            col=RPAD_COL,
            row=DATA_FIRST_ROW),
        TYPE_HERE_STYLE)

    fill_cell(
        sheet,
        HEADER_ROW,
        DATA_FIRST_COL_NUM,
        recombinant_language_text(chromo['title'])
            + u' \N{em dash} ' + org_title_lang_hack(org['title']),
        'reco_header')

    sheet.cell(row=CODE_ROW, column=1).value = 'v3'  # template version
    # allow only upload to this org
    sheet.cell(row=CODE_ROW, column=2).value = org['name']

    cheadings_dimensions = sheet.row_dimensions[CHEADINGS_ROW]

    choice_fields = dict(
        (f['datastore_id'], f['choices'])
        for f in recombinant_choice_fields(chromo['resource_name']))

    for col_num, field in template_cols_fields(chromo):
        field_heading = recombinant_language_text(
            field.get('excel_heading', field['label'])).strip()
        cheadings_dimensions.height = max(
            cheadings_dimensions.height,
            field_heading.count('\n') * LINE_HEIGHT + CHEADINGS_HEIGHT)

        col_heading_style = 'reco_cheading'
        if 'excel_column_heading_style' in field:
            # use geno column heading style as base, just override keys
            col_heading_style = dict(
                dict(
                    DEFAULT_CHEADING_STYLE,
                    **geno.get('excel_column_heading_style', {})),
                **field['excel_column_heading_style'])
            apply_style(sheet.cell(
                row=HEADER_ROW, column=col_num), col_heading_style)
            apply_style(sheet.cell(
                row=CSTATUS_ROW, column=col_num), col_heading_style)

        fill_cell(
            sheet,
            CHEADINGS_ROW,
            col_num,
            field_heading,
            col_heading_style)

        reference_row1 = len(refs) + REF_FIRST_ROW

        # match against db columns
        sheet.cell(row=CODE_ROW, column=col_num).value = field['datastore_id']

        example = chromo['examples']['record'].get(field['datastore_id'], '')
        fill_cell(
            sheet,
            EXAMPLE_ROW,
            col_num,
            u','.join(example) if isinstance(example, list) else example,
            'reco_example')

        col_letter = get_column_letter(col_num)

        # jump to first error/required cell in column
        fill_cell(
            sheet,
            CSTATUS_ROW,
            col_num,
            '=IF(e{rnum}!{col}{row}>0,HYPERLINK("#{col}"&e{rnum}!{col}{row},"")'
                ',IF(r{rnum}!{col}{row}>0,HYPERLINK("#{col}"&r{rnum}!{col}{row},""),""))'
                .format(rnum=resource_num, col=col_letter, row=CSTATUS_ROW),
            col_heading_style)

        col = sheet.column_dimensions[col_letter]
        if 'excel_column_width' in field:
            col.width = field['excel_column_width']
        else:
            col.width = max(estimate_width(field_heading), CHEADINGS_MIN_WIDTH)

        validation_range = '{col}{row1}:{col}{rowN}'.format(
            col=col_letter,
            row1=DATA_FIRST_ROW,
            rowN=DATA_FIRST_ROW + data_num_rows - 1)

        xl_format = datastore_type[field['datastore_type']].xl_format
        alignment = openpyxl.styles.Alignment(wrap_text=True)
        col_style = NamedStyle(
            name='reco_{0}{1}'.format(resource_num, col_letter),
            number_format=xl_format,
            alignment=alignment,
            protection=openpyxl.styles.Protection(locked=False))
        book.add_named_style(col_style)
        for (c,) in sheet[validation_range]:
            c.style = col_style.name
        ex_cell = sheet.cell(row=EXAMPLE_ROW, column=col_num)
        ex_cell.number_format = xl_format
        ex_cell.alignment = alignment

        _append_field_ref_rows(refs, field, "#'{sheet}'!{col}{row}".format(
            sheet=sheet.title, col=col_letter, row=CHEADINGS_ROW))

        if field['datastore_id'] in choice_fields:
            full_text_choices = (
                field['datastore_type'] != '_text' and field.get(
                'excel_full_text_choices', False))
            ref1 = len(refs) + REF_FIRST_ROW
            max_choice_width = _append_field_choices_rows(
                refs,
                choice_fields[field['datastore_id']],
                full_text_choices)
            refN = len(refs) + REF_FIRST_ROW - 2

            if full_text_choices:
                if 'excel_column_width' not in field:
                    col.width = max(col.width, max_choice_width)
                # expand example
                for ck, cv in choice_fields[field['datastore_id']]:
                    if ck == example:
                        ex_cell.value = u"{0}: {1}".format(ck, cv)
                        break

            choice_range = 'reference!${col}${ref1}:${col}${refN}'.format(
                col=REF_KEY_COL, ref1=ref1, refN=refN)
            user_choice_range = field.get('excel_choice_range_formula')
            if user_choice_range:
                choice_keys = set(
                    key for (_i, key, _i, _i) in string.Formatter().parse(user_choice_range)
                    if key != 'range' and key != 'range_top')
                choice_values = {}
                if choice_keys:
                    choice_values = {
                        f['datastore_id']: "{col}{num}".format(
                            col=get_column_letter(cn),
                            num=DATA_FIRST_ROW)
                        for cn, f in template_cols_fields(chromo)
                        if f['datastore_id'] in choice_keys}
                user_choice_range = user_choice_range.format(
                    range=choice_range,
                    range_top=choice_range.split(':')[0],
                    **choice_values)
            cranges[field['datastore_id']] = choice_range

            choices = [c[0] for c in choice_fields[field['datastore_id']]]
            if field['datastore_type'] != '_text':
                v = openpyxl.worksheet.datavalidation.DataValidation(
                    type="list",
                    formula1=user_choice_range or choice_range,
                    allow_blank=True)
                v.errorTitle = u'Invalid choice'
                valid_keys = u', '.join(unicode(c) for c in choices)
                if len(valid_keys) < 40:
                    v.error = (u'Please enter one of the valid keys: '
                        + valid_keys)
                else:
                    v.error = (u'Please enter one of the valid keys shown on '
                        'sheet "reference" rows {0}-{1}'.format(ref1, refN))
                sheet.add_data_validation(v)
                v.add(validation_range)

        sheet.cell(row=CHEADINGS_ROW, column=col_num).hyperlink = (
            '#reference!{colA}{row1}:{colZ}{rowN}'.format(
                colA=REF_FIELD_NUM_COL,
                row1=reference_row1,
                colZ=REF_VALUE_COL,
                rowN=len(refs) + REF_FIRST_ROW - 2))

    _add_conditional_formatting(
        sheet,
        col_letter,
        resource_num,
        error_style,
        required_style,
        data_num_rows)

    sheet.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT
    sheet.row_dimensions[CODE_ROW].hidden = True
    sheet.row_dimensions[CSTATUS_ROW].height = CSTATUS_HEIGHT
    sheet.row_dimensions[EXAMPLE_ROW].height = chromo.get(
        'excel_example_height', DEFAULT_EXAMPLE_HEIGHT)
    for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + data_num_rows):
        sheet.row_dimensions[i].height = chromo.get(
            'excel_data_height', DEFAULT_DATA_HEIGHT)

        # jump to first error/required cell in row
        sheet.cell(row=i, column=RSTATUS_COL_NUM).value = (
            '=IF(e{rnum}!{col}{row}>0,'
                'HYPERLINK("#"&ADDRESS({row},e{rnum}!{col}{row}),""),'
                'IF(r{rnum}!{col}{row}>0,'
                    'HYPERLINK("#"&ADDRESS({row},r{rnum}!{col}{row}),""),""))'
            .format(rnum=resource_num, col=RSTATUS_COL, row=i))

    sheet.column_dimensions[RSTATUS_COL].width = RSTATUS_WIDTH
    sheet.column_dimensions[RPAD_COL].width = RPAD_WIDTH

    sheet.freeze_panes = sheet[FREEZE_PANES]

    apply_style(sheet.row_dimensions[HEADER_ROW], header_style)
    apply_style(sheet.row_dimensions[CHEADINGS_ROW], cheadings_style)
    apply_style(sheet.row_dimensions[CSTATUS_ROW], cheadings_style)
    apply_style(sheet.row_dimensions[EXAMPLE_ROW], example_style)
    for (c,) in sheet[EDGE_RANGE]:
        c.style = 'reco_edge'

    # trying to set the active cell (not working yet)
    select = "{col}{row}".format(col=DATA_FIRST_COL, row=DATA_FIRST_ROW)
    sheet.sheet_view.selection[0].activeCell = select
    sheet.sheet_view.selection[0].sqref = select

    return cranges


def _append_field_ref_rows(refs, field, link):
    refs.append((None, []))
    label = recombinant_language_text(field['label'])
    refs.append(('title', [(link, label) if link else label]))
    refs.append(('attr', [
        _('ID'),
        field['datastore_id']]))
    if 'description' in field:
        refs.append(('attr', [
            _('Description'),
            recombinant_language_text(field['description'])]))
    if 'obligation' in field:  # for old yaml files (merged with validation in new ones)
        refs.append(('attr', [
            _('Obligation'),
            recombinant_language_text(field['obligation'])]))
    if 'validation' in field:
        refs.append(('attr', [
            _('Validation'),
            recombinant_language_text(field['validation'])]))
    if 'format_type' in field:  # for old yaml files (merged with validation in new ones)
        refs.append(('attr', [
            _('Format'),
            recombinant_language_text(field['format_type'])]))

def _append_field_choices_rows(refs, choices, full_text_choices):
    refs.append(('choice heading', [_('Values')]))
    max_length = 0
    for key, value in choices:
        if full_text_choices:
            choice = [u'{0}: {1}'.format(key, value)]
        elif unicode(key) == value:
            choice = [unicode(key)]
        else:
            choice = [unicode(key), value]
        refs.append(('choice', choice))
        max_length = max(max_length, len(choice[0]))  # used for full_text_choices
    return estimate_width_from_length(max_length)

def _populate_reference_sheet(sheet, geno, refs):
    field_count = 1

    header1_style = dict(DEFAULT_HEADER_STYLE, **geno.get('excel_header_style', {}))
    header2_style = dict(DEFAULT_REF_HEADER2_STYLE, **geno.get('excel_header_style', {}))
    fill_cell(
        sheet,
        REF_HEADER1_ROW,
        REF_KEY_COL_NUM,
        recombinant_language_text(geno['title']),
        'reco_header')
    apply_style(sheet.row_dimensions[REF_HEADER1_ROW], header1_style)
    fill_cell(
        sheet,
        REF_HEADER2_ROW,
        REF_KEY_COL_NUM,
        _('Reference'),
        'reco_header2')
    apply_style(sheet.row_dimensions[REF_HEADER2_ROW], header2_style)
    for (c,) in sheet[REF_EDGE_RANGE]:
        c.style = 'reco_edge'
    sheet.row_dimensions[REF_HEADER1_ROW].height = REF_HEADER1_HEIGHT
    sheet.row_dimensions[REF_HEADER2_ROW].height = REF_HEADER2_HEIGHT


    for row_number, (style, ref_line) in enumerate(refs, REF_FIRST_ROW - 1):
        link = None
        if len(ref_line) == 2:
            value = wrap_text_to_width(ref_line[1], REF_VALUE_WIDTH).strip()
            ref_line = [ref_line[0], value]
        elif len(ref_line) == 1 and isinstance(ref_line[0], tuple):
            link, value = ref_line[0]
            value = value.strip()
            ref_line = [value]

        for cnum, cval in enumerate(ref_line, REF_KEY_COL_NUM):
            sheet.cell(row=row_number, column=cnum).value = (
                cval.strip().replace('\n', '\r\n'))

        if len(ref_line) == 2:
            sheet.row_dimensions[row_number].height = LINE_HEIGHT + (
                value.count('\n') * LINE_HEIGHT)

        key_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM)
        value_cell = sheet.cell(row=row_number, column=REF_VALUE_COL_NUM)

        if style == 'title':
            sheet.merge_cells(REF_FIELD_NUM_MERGE.format(row=row_number))
            sheet.merge_cells(REF_FIELD_TITLE_MERGE.format(row=row_number))
            fill_cell(
                sheet,
                row_number,
                REF_FIELD_NUM_COL_NUM,
                field_count,
                'reco_ref_number')
            title_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM)
            if link:
                title_cell.hyperlink = link
            title_cell.style = 'reco_ref_title'
            sheet.row_dimensions[row_number].height = REF_FIELD_TITLE_HEIGHT
            field_count += 1
        elif style == 'choice':
            pad_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM - 1)
            pad_cell.style = 'reco_example'
            key_cell.style = 'reco_example'
            value_cell.style = 'reco_example'
        elif style == 'attr':
            key_cell.style = 'reco_ref_attr'
            value_cell.style = 'reco_ref_value'
        elif style == 'choice heading':
            key_cell.style = 'reco_ref_attr'
            value_cell.style = 'reco_ref_value'
            sheet.row_dimensions[row_number].height = REF_CHOICE_HEADING_HEIGHT

        apply_style(sheet.row_dimensions[row_number], REF_PAPER_STYLE)

    sheet.column_dimensions[RSTATUS_COL].width = RSTATUS_WIDTH
    sheet.cell(row=1, column=RPAD_COL_NUM).value = None  # make sure rpad col exists
    sheet.column_dimensions[RPAD_COL].width = RPAD_WIDTH
    sheet.column_dimensions[REF_KEY_COL].width = REF_KEY_WIDTH
    sheet.column_dimensions[REF_VALUE_COL].width = REF_VALUE_WIDTH


def _populate_excel_e_sheet(sheet, chromo, cranges):
    """
    Populate the "error" calculation excel worksheet

    The 'A' column is the sum of all following columns.
    The 4th row is the sum of all rows below.

    Other cells are 1 for error, 0 or blank for no error or no value
    in the corresponding cell on the data entry sheet.
    """
    col = None
    data_num_rows = chromo.get('excel_data_num_rows', DEFAULT_DATA_NUM_ROWS)

    for col_num, field in template_cols_fields(chromo):
        pk_field = field['datastore_id'] in chromo['datastore_primary_key']

        crange = cranges.get(field['datastore_id'])
        fmla = None
        if field['datastore_type'] == 'date':
            fmla = 'NOT(ISNUMBER({cell}+0))'
        elif field['datastore_type'] == 'int':
            fmla = 'NOT(IFERROR(INT({cell})=VALUE({cell}),FALSE))'
        elif field['datastore_type'] == 'year':
            fmla = (
                'NOT(IFERROR(AND(INT({{cell}})={{cell}},'
                '{{cell}}>={year_min},{{cell}}<={year_max}),FALSE))'
                ).format(
                    year_min=chromo.get('year_min', DEFAULT_YEAR_MIN),
                    year_max=chromo.get('year_max', DEFAULT_YEAR_MAX))
        elif field['datastore_type'] == 'numeric':
            fmla = 'NOT(ISNUMBER({cell}))'
        elif field['datastore_type'] == 'money':
            fmla = (
                'NOT(IFERROR(ROUND(VALUE({cell}),2)=VALUE({cell}),FALSE))')
        elif crange and field['datastore_type'] == '_text':
            # multiple comma-separated choices
            # validate by counting choices against matches
            fmla = (
                'LEN(SUBSTITUTE({{cell}}," ",""))+1-SUMPRODUCT(--ISNUMBER('
                'SEARCH(","&{r}&",",SUBSTITUTE(","&{{cell}}&","," ",""))),'
                'LEN({r})+1)').format(r=crange)
        elif crange and field.get('excel_full_text_choices', False):
            # 'code:text'-style choices, accept 'code' and 'code:anything'
            fmla = (
                'COUNTIF({r},TRIM(LEFT({{cell}},FIND(":",{{cell}}&":")-1))&":*")=0'
                ).format(r=crange)
        elif crange:
            # single choice
            fmla = 'COUNTIF({r},TRIM({{cell}}))=0'.format(r=crange)

        user_fmla = field.get('excel_error_formula')
        if user_fmla:
            if not fmla:
                fmla = 'FALSE()'
            fmla = user_fmla.replace('{default_formula}', '(' + fmla + ')')

        if field.get('excel_error_cell_filter_formula'):
            fmla = fmla.replace(
                '{cell}', '(' + field.get('excel_error_cell_filter_formula') + ')')

        if pk_field:
            # repeated primary (composite) keys are errors
            pk_fmla = 'SUMPRODUCT(' + ','.join(
                "--(TRIM('{sheet}'!{col}{top}:{col}{{num}})"
                "=TRIM('{sheet}'!{col}{{num}}))".format(
                    sheet=chromo['resource_name'],
                    col=get_column_letter(cn),
                    top=DATA_FIRST_ROW)
                for cn, f in template_cols_fields(chromo)
                if f['datastore_id'] in chromo['datastore_primary_key']
                ) +')>1'
            fmla = ('OR(' + fmla + ',' + pk_fmla + ')') if fmla else pk_fmla

        if not fmla:
            continue

        fmla_keys = set(
            key for (_i, key, _i, _i) in string.Formatter().parse(fmla)
            if key != 'cell' and key != 'default_formula')
        if fmla_keys:
            fmla_values = {
                f['datastore_id']: "'{sheet}'!{col}{{num}}".format(
                    sheet=chromo['resource_name'],
                    col=get_column_letter(cn))
                for cn, f in template_cols_fields(chromo)
                if f['datastore_id'] in fmla_keys}

        col = get_column_letter(col_num)
        cell = "'{sheet}'!{col}{{num}}".format(
            sheet=chromo['resource_name'],
            col=col)
        fmla = '=NOT({cell}="")*(' + fmla + ')'
        for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + data_num_rows):
            try:
                sheet.cell(row=i, column=col_num).value = fmla.format(
                    cell=cell,
                    num='{num}',
                    **fmla_values).format(num=i)
            except KeyError:
                assert 0, (fmla, fmla_values)

        sheet.cell(row=CSTATUS_ROW, column=col_num).value = (
            '=IFERROR(MATCH(TRUE,INDEX({col}{row1}:{col}{rowN}<>0,),)+{row0},0)'
            .format(
                col=col,
                row1=DATA_FIRST_ROW,
                row0=DATA_FIRST_ROW - 1,
                rowN=DATA_FIRST_ROW + data_num_rows - 1))

    if col is None:
        return  # no errors to report on!

    for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + data_num_rows):
        sheet.cell(row=i, column=RSTATUS_COL_NUM).value = (
            '=IFERROR(MATCH(TRUE,INDEX({colA}{row}:{colZ}{row}<>0,),)+{col0},0)'.format(
                colA=DATA_FIRST_COL,
                col0=DATA_FIRST_COL_NUM - 1,
                colZ=col,
                row=i))


def _populate_excel_r_sheet(sheet, chromo):
    """
    Populate the "required" calculation excel worksheet

    The 'A' column is the sum of all columns "C" and later.
    The 'B' column is TRUE when any data is entered on the corresponding row
    of the data entry sheet.
    The 4th row is the sum of all rows below.

    Other cells in this worksheet are 1 for required fields, 0 or blank for
    no value or not required fields in the corresponding cell on the
    data entry sheet
    """
    col = None
    data_num_rows = chromo.get('excel_data_num_rows', DEFAULT_DATA_NUM_ROWS)

    for col_num, field in template_cols_fields(chromo):
        fmla = field.get('excel_required_formula')
        pk_field = field['datastore_id'] in chromo['datastore_primary_key']

        if fmla:
            fmla = '={has_data}*({cell}="")*(' + fmla +')'
        elif pk_field or field.get('excel_required', False):
            fmla = '={has_data}*({cell}="")'
        else:
            continue

        col = get_column_letter(col_num)
        cell = "'{sheet}'!{col}{{num}}".format(
            sheet=chromo['resource_name'],
            col=col)

        fmla_keys = set(
            key for (_i, key, _i, _i) in string.Formatter().parse(fmla)
            if key != 'cell' and key != 'has_data')
        fmla_values = {}
        if fmla_keys:
            fmla_values = {
                f['datastore_id']: "'{sheet}'!{col}{{num}}".format(
                    sheet=chromo['resource_name'],
                    col=get_column_letter(cn))
                for cn, f in template_cols_fields(chromo)
                if f['datastore_id'] in fmla_keys}

        for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + data_num_rows):
            sheet.cell(row=i, column=col_num).value = fmla.format(
                cell=cell,
                has_data='{col}{{num}}'.format(col=RPAD_COL),
                **fmla_values).format(num=i)

        sheet.cell(row=CSTATUS_ROW, column=col_num).value = (
            '=IFERROR(MATCH(TRUE,INDEX({col}{row1}:{col}{rowN}<>0,),)+{row0},0)'
            .format(
                col=col,
                row1=DATA_FIRST_ROW,
                row0=DATA_FIRST_ROW - 1,
                rowN=DATA_FIRST_ROW + data_num_rows - 1))

    if col is None:
        return  # no required columns

    for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + data_num_rows):
        sheet.cell(row=i, column=RPAD_COL_NUM).value = (
            "=SUMPRODUCT(LEN('{sheet}'!{colA}{row}:{colZ}{row}))>0".format(
                sheet=chromo['resource_name'],
                colA=DATA_FIRST_COL,
                colZ=col,
                row=i))

    for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + data_num_rows):
        sheet.cell(row=i, column=RSTATUS_COL_NUM).value = (
            '=IFERROR(MATCH(TRUE,INDEX({colA}{row}:{colZ}{row}<>0,),)+{col0},0)'
            .format(
                colA=DATA_FIRST_COL,
                col0=DATA_FIRST_COL_NUM - 1,
                colZ=col,
                row=i))

def fill_cell(sheet, row, column, value, style):
    """
    :param sheet: worksheet
    :param row: 1-based row number
    :param column: 1-based column number
    :param value: value to store (unicode, int, date, ..)
    :param style: style name as string or dict for apply_style
    :return: None
    """
    c = sheet.cell(row=row, column=column)
    if isinstance(value, basestring):
        value = value.replace(u'\n', u'\r\n')
    c.value = value
    if isinstance(style, basestring):
        c.style = style
    else:
        apply_style(c, style)


def build_named_style(book, name, config):
    """
    :param book: workbook to assign style
    :param name: style name
    :param config: dict with style configuration
    :return: None
    """
    style = NamedStyle(name=name)
    apply_style(style, config)
    book.add_named_style(style)


def apply_style(target, config):
    """
    apply style from config to target

    currently supports PatternFill, Font, Alignment
    :param target: object to assign
    :param config:
    """
    pattern_fill = config.get('PatternFill')
    if pattern_fill:
        target.fill = openpyxl.styles.PatternFill(**pattern_fill)
    font = config.get('Font')
    if font:
        target.font = openpyxl.styles.Font(**font)
    alignment = config.get('Alignment')
    if alignment:
        target.alignment = openpyxl.styles.Alignment(**alignment)


def org_title_lang_hack(title):
    """
    Canada site is using title to store "{en title name} | {fr title name}"
    this hack displays the correct one (one day soon we'll fix this, promise)
    """
    try:
        lang = h.lang()
    except TypeError:
        lang = 'en'
    if lang == 'fr':
        return title.split(u' | ')[-1]
    return title.split(u' | ')[0]

def template_cols_fields(chromo):
    ''' (col_num, field) ... for fields in template'''
    return enumerate(
        (f for f in chromo['fields'] if f.get(
            'import_template_include', True)), DATA_FIRST_COL_NUM)

def _add_conditional_formatting(
        sheet, col_letter, resource_num, error_style, required_style,
        data_num_rows):
    '''
    Error and required cell hilighting based on e/r sheets
    '''
    error_fill = openpyxl.styles.PatternFill(
        bgColor=error_style['PatternFill']['fgColor'],
        **error_style['PatternFill'])
    error_font = openpyxl.styles.Font(**error_style['Font'])
    required_fill = openpyxl.styles.PatternFill(
        bgColor=required_style['PatternFill']['fgColor'],
        **required_style['PatternFill'])
    required_font = openpyxl.styles.Font(**required_style['Font'])

    sheet.conditional_formatting.add(
        '{col}{row1}:{col}{rowN}'.format(
            col=RSTATUS_COL,
            row1=DATA_FIRST_ROW,
            rowN=DATA_FIRST_ROW + data_num_rows - 1),
        FormulaRule([
            'AND(e{rnum}!{colA}{row1}=0,r{rnum}!{colA}{row1}>0)'.format(
                rnum=resource_num,
                colA=RSTATUS_COL,
                row1=DATA_FIRST_ROW)],
        stopIfTrue=True,
        fill=required_fill,
        font=required_font))
    sheet.conditional_formatting.add(
        '{colA}{row1}:{colZ}{rowN}'.format(
            colA=RSTATUS_COL,
            row1=CSTATUS_ROW,
            colZ=col_letter,
            rowN=DATA_FIRST_ROW + data_num_rows - 1),
        FormulaRule([
            'AND(ISNUMBER(e{rnum}!{colA}{row1}),'
            'e{rnum}!{colA}{row1}>0)'.format(
                rnum=resource_num,
                colA=RSTATUS_COL,
                row1=CSTATUS_ROW)],
        stopIfTrue=True,
        fill=error_fill,
        font=error_font))
    sheet.conditional_formatting.add(
        '{colA}{row1}:{colZ}{rowN}'.format(
            colA=DATA_FIRST_COL,
            row1=CSTATUS_ROW,
            colZ=col_letter,
            rowN=DATA_FIRST_ROW + data_num_rows - 1),
        FormulaRule([
            'AND(ISNUMBER(r{rnum}!{colA}{row1}),'
            'e{rnum}!{colA}{row1}=0,r{rnum}!{colA}{row1}>0)'.format(
                rnum=resource_num,
                colA=DATA_FIRST_COL,
                row1=CSTATUS_ROW)],
        stopIfTrue=True,
        fill=required_fill,
        font=required_font))
