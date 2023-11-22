"""
Old excel v2 template code, remove when no longer used
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

from ckanext.recombinant.tables import get_geno
from ckanext.recombinant.errors import RecombinantException
from ckanext.recombinant.datatypes import datastore_type
from ckanext.recombinant.helpers import (
    recombinant_choice_fields, recombinant_language_text)

from ckan.plugins.toolkit import _

white_font = openpyxl.styles.Font(color=openpyxl.styles.colors.WHITE)


def _populate_excel_sheet_v2(sheet, chromo, org, refs):
    """
    Format openpyxl sheet for the resource definition chromo and org.

    refs - list of rows to add to reference sheet, modified
        in place from this function

    returns field information for reference sheet
    """
    sheet.title = chromo['resource_name']

    org_style = dict(
        chromo['excel_organization_style'],
        Alignment={'vertical': 'center'})
    fill_cell(sheet, 1, 1, org['name'], org_style)
    fill_cell(
        sheet,
        1,
        2,
        recombinant_language_text(chromo['title']) + '        ' + org['title'],
        org_style)
    sheet.row_dimensions[1].height = 24
    apply_styles(org_style, sheet.row_dimensions[1])

    header_style = chromo['excel_header_style']
    error_color = chromo.get('excel_error_background_color', '763626')
    required_color = chromo.get('excel_required_border_color', '763626')

    error_fill = openpyxl.styles.PatternFill(
        start_color='FF%s' % error_color,
        end_color='FF%s' % error_color,
        fill_type='solid')
    required_side = openpyxl.styles.Side(
        style='medium',
        color='FF%s' % required_color)
    required_border = openpyxl.styles.Border(
        required_side, required_side, required_side, required_side)


    choice_fields = dict(
        (f['datastore_id'], f['choices'])
        for f in recombinant_choice_fields(chromo['resource_name']))

    pk_cells = [
        get_column_letter(n)+'4' for
        n, field in enumerate((f for f in chromo['fields'] if f.get(
                    'import_template_include', True)), 1)
        if field['datastore_id'] in chromo['datastore_primary_key']]

    for n, field in enumerate((f for f in chromo['fields'] if f.get(
            'import_template_include', True)), 1):
        fill_cell(sheet, 2, n, recombinant_language_text(field['label']), header_style)
        fill_cell(sheet, 3, n, field['datastore_id'], header_style)
        # jumping through openpyxl hoops:
        col_letter = get_column_letter(n)
        col_letter_before = get_column_letter(max(1, n-1))
        col_letter_after = get_column_letter(n+1)
        col = sheet.column_dimensions[col_letter]
        col.width = field['excel_column_width']
        col.alignment = openpyxl.styles.Alignment(
            wrap_text=True)
        # FIXME: format only below header
        col.number_format = datastore_type[field['datastore_type']].xl_format
        validation_range = '{0}4:{0}1004'.format(col_letter)

        _append_field_ref_rows_v2(refs, field, org_style, header_style)

        if field['datastore_type'] == 'boolean':
            v = openpyxl.worksheet.datavalidation.DataValidation(
                type="list", formula1='"FALSE,TRUE"', allow_blank=True)
            sheet.add_data_validation(v)
            v.add(validation_range)
        if field['datastore_type'] == 'date':
            sheet.conditional_formatting.add(validation_range,
                FormulaRule([
                        # +0 is needed by excel to recognize dates. sometimes.
                        'AND(NOT(ISBLANK({cell})),NOT(ISNUMBER({cell}+0)))'
                        .format(cell=col_letter + '4',)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                FormulaRule([
                        # +0 is needed by excel to recognize dates. sometimes.
                        'SUMPRODUCT(--NOT(ISBLANK({cells})),'
                        '--NOT(ISNUMBER({cells}+0)))'
                        .format(cells=validation_range,)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
        if field['datastore_type'] == 'int':
            sheet.conditional_formatting.add(validation_range,
                FormulaRule([
                        'AND(NOT(ISBLANK({cell})),NOT(IFERROR(INT({cell})={cell},FALSE)))'
                        .format(cell=col_letter + '4',)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                FormulaRule([
                        'SUMPRODUCT(--NOT(ISBLANK({cells})),'
                        '--NOT(IFERROR(INT({cells})={cells},FALSE)))'
                        .format(cells=validation_range,)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
        if field['datastore_type'] == 'money':
            sheet.conditional_formatting.add(validation_range,
                FormulaRule([
                        # isblank doesnt work. sometimes. trim()="" is more permissive
                        'AND(NOT(TRIM({cell})=""),NOT(IFERROR(ROUND({cell},2)={cell},FALSE)))'
                        .format(cell=col_letter + '4',)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                FormulaRule([
                        # isblank doesnt work. sometimes. trim()="" is more permissive
                        'SUMPRODUCT(--NOT(TRIM({cells})=""),'
                        '--NOT(IFERROR(ROUND({cells},2)={cells},FALSE)))'
                        .format(cells=validation_range,)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))


        if field['datastore_id'] in choice_fields:
            ref1 = len(refs) + 1
            _append_field_choices_rows_v2(
                refs,
                choice_fields[field['datastore_id']],
                header_style,
                sheet.title + '!' + validation_range
                if field['datastore_type'] == '_text' else None)
            refN = len(refs)

            choice_range = 'reference!$B${0}:$B${1}'.format(ref1, refN)

            choices = [c[0] for c in choice_fields[field['datastore_id']]]
            if field['datastore_type'] == '_text':
                # custom validation only works in Excel, use conditional
                # formatting for libre office compatibility
                sheet.conditional_formatting.add(validation_range,
                    FormulaRule([(
                        # count characters in the cell
                        'IF(SUBSTITUTE({col}4," ","")="",0,'
                        'LEN(SUBSTITUTE({col}4," ",""))+1)-'
                        # minus length of valid choices
                        'SUMPRODUCT(--ISNUMBER(SEARCH('
                        '","&{r}&",",SUBSTITUTE(","&{col}4&","," ",""))),'
                        'LEN({r})+1)'
                        .format(
                            col=col_letter,
                            r=choice_range)
                        )],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            else:
                v = openpyxl.worksheet.datavalidation.DataValidation(
                    type="list",
                    formula1=choice_range,
                    allow_blank=True)
                v.errorTitle = u'Invalid choice'
                valid_keys = u', '.join(str(c) for c in choices)
                if len(valid_keys) < 40:
                    v.error = (u'Please enter one of the valid keys: '
                        + valid_keys)
                else:
                    v.error = (u'Please enter one of the valid keys shown on '
                        'sheet "reference" rows {0}-{1}'.format(ref1, refN))
                sheet.add_data_validation(v)
                v.add(validation_range)

            # hilight header if bad values pasted below
            if field['datastore_type'] == '_text':
                choice_counts = 'reference!$J${0}:$J${1}'.format(ref1, refN)
                sheet.conditional_formatting.add("{0}2".format(col_letter),
                    FormulaRule([(
                            # count characters in the validation range
                            'SUMPRODUCT(IF(SUBSTITUTE({v}," ","")="",0,'
                            'LEN(SUBSTITUTE({v}," ",""))+1))-'
                            # minus length of all valid choices found
                            'SUMPRODUCT({counts},LEN({choices})+1)'
                            .format(
                                v=validation_range,
                                col=col_letter,
                                choices=choice_range,
                                counts=choice_counts)
                            )],
                        stopIfTrue=True,
                        fill=error_fill,
                        font=white_font,
                        ))
            else:
                sheet.conditional_formatting.add("{0}2".format(col_letter),
                    FormulaRule([(
                            'SUMPRODUCT(--NOT(TRIM({0})=""))'
                            '-SUMPRODUCT(COUNTIF({1},TRIM({0})))'
                            .format(validation_range, choice_range))],
                        stopIfTrue=True,
                        fill=error_fill,
                        font=white_font,
                        ))

        if field.get('excel_cell_required_formula'):
            sheet.conditional_formatting.add(validation_range,
                FormulaRule([
                        field['excel_cell_required_formula'].format(
                            column=col_letter,
                            column_before=col_letter_before,
                            column_after=col_letter_after,
                            row='4',
                        )],
                    stopIfTrue=True,
                    border=required_border,
                    ))
        elif (field.get('excel_required') or
                field['datastore_id'] in chromo['datastore_primary_key']):
            # hilight missing values
            sheet.conditional_formatting.add(validation_range,
                FormulaRule([(
                        'AND({col}4="",SUMPRODUCT(LEN(A4:Z4)))'
                        .format(col=col_letter)
                        )],
                    stopIfTrue=True,
                    border=required_border,
                    ))
        if field.get('excel_cell_error_formula'):
            sheet.conditional_formatting.add(validation_range,
                FormulaRule([
                    field['excel_cell_error_formula'].format(
                        cell=col_letter + '4',)
                    ],
                stopIfTrue=True,
                fill=error_fill,
                font=white_font,
                ))
        if field.get('excel_header_error_formula'):
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                FormulaRule([
                        field['excel_header_error_formula'].format(
                            cells=validation_range,
                            column=col_letter,
                        )],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))

    apply_styles(header_style, sheet.row_dimensions[2])
    apply_styles(header_style, sheet.row_dimensions[3])
    sheet.row_dimensions[3].hidden = True

    sheet.freeze_panes = sheet['A4']


def _append_field_ref_rows_v2(refs, field, style1, style2):
    a1 = (style2, style1, 24)
    a2 = (style2, None, None)
    refs.append((None, []))
    refs.append((a1, [
        _('Field Name'),
        recombinant_language_text(field['label'])]))
    refs.append((a2, [
        _('ID'),
        field['datastore_id']]))
    if 'description' in field:
        refs.append((a2, [
            _('Description'),
            recombinant_language_text(field['description'])]))
    if 'obligation' in field:
        refs.append((a2, [
            _('Obligation'),
            recombinant_language_text(field['obligation'])]))
    if 'format_type' in field:
        refs.append((a2, [
            _('Format'),
            recombinant_language_text(field['format_type'])]))

def _append_field_choices_rows_v2(refs, choices, style2, count_range=None):
    label = _('Values')
    a1 = (style2, None, 24)
    for key, value in choices:
        r = [label, str(key), value]
        if count_range: # used by _text choices validation
            r.extend([None]*6 + ['=SUMPRODUCT(--ISNUMBER(SEARCH('
                '","&B{n}&",",SUBSTITUTE(","&{r}&","," ",""))))'.format(
                    r=count_range,
                    n=len(refs) + 1)])
        refs.append((a1, r))
        label = None
        a1 = (style2, None, None)


def _populate_reference_sheet_v2(sheet, chromo, refs):
    for style, ref_line in refs:
        sheet.append(ref_line)
        if not style:
            continue

        s1, s2, height = style
        if height:
            sheet.row_dimensions[sheet.max_row].height = height

        if s2:
            apply_styles(s2, sheet.row_dimensions[sheet.max_row])
        for c in range(len(ref_line)):
            if c and s2:
                apply_styles(s2, sheet.cell(
                    row=sheet.max_row, column=c + 1))
            if not c and s1:
                apply_styles(s1, sheet.cell(
                    row=sheet.max_row, column=c + 1))


def fill_cell(sheet, row, column, value, styles):
    c = sheet.cell(row=row, column=column)
    if isinstance(value, str):
        value = value.replace(u'\n', u'\r\n')
    c.value = value
    apply_styles(styles, c)


def apply_styles(config, target):
    """
    apply styles from config to target

    currently supports PatternFill, Font, Alignment
    """
    pattern_fill = config.get('PatternFill')
    if pattern_fill:
        target.fill = openpyxl.styles.PatternFill(**pattern_fill)
    font = config.get('Font')
    if font:
        target.font = openpyxl.styles.Font(**font)
    alignment = config.get('Alignment')
    if alignment:
        target.alignment = openpyxl.styles.Alignment(**alignment)
