import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from typing import Any, List, Dict, Iterator, Union, Optional, Tuple

from werkzeug.datastructures import FileStorage as FlaskFileStorage
from cgi import FieldStorage

from ckanext.recombinant.datatypes import canonicalize
from ckanext.recombinant.errors import BadExcelData


HEADER_ROWS_V2 = 3
HEADER_ROWS_V3 = 5


def read_excel(f: Union[str, FlaskFileStorage, FieldStorage],
               file_contents: Optional[str] = None) -> Iterator[Any]:
    """
    Return a generator that opens the excel file f (name or file object)
    and then produces ((sheet-name, org-name), row1, row2, ...)
    :param: f: file name or xlsx file object

    :return: Generator that opens the excel file f
    and then produces:
        (sheet-name, org-name, column_names, data_rows_generator)
        ...
    :rtype: generator
    """
    # type_ignore_reason: incomplete typing
    wb = load_workbook(f, read_only=True)  # type: ignore

    for sheetname in wb.sheetnames:
        if sheetname == 'reference':
            return
        # type_ignore_reason: incomplete typing
        sheet: Worksheet = wb[sheetname]
        rowiter = sheet.rows
        # type_ignore_reason: incomplete typing
        organization_row = next(rowiter)

        # type_ignore_reason: incomplete typing
        next(rowiter)
        names_row = next(rowiter)

        org_name = organization_row[0].value
        if org_name and names_row[0].value != 'v3':
            # v2 template
            yield (
                sheetname,
                org_name,
                [c.value for c in names_row],
                # type_ignore_reason: incomplete typing
                _filter_bumf(rowiter, HEADER_ROWS_V2))
            continue

        # type_ignore_reason: incomplete typing
        next(rowiter)
        example_row = next(rowiter)
        if example_row[0].value != 'e.g.' and example_row[0].value != 'ex.':
            raise BadExcelData('Example record on row 5 is missing')

        yield (
            sheetname,
            names_row[1].value,
            [c.value for c in names_row[2:]],
            _filter_bumf((row[2:] for row in rowiter), HEADER_ROWS_V3))


def _filter_bumf(rowiter: Iterator[Any],
                 header_rows: int) -> Iterator[Any]:
    i = header_rows
    for row in rowiter:
        i += 1
        values = [
            unescape(c.value) if isinstance(c.value, str) else c.value
            for c in row]
        # return next non-empty row
        if not all(_is_bumf(v) for v in values):
            yield i, values


def _is_bumf(value: Any) -> Union[str, bool]:
    """
    Return true if this value is filler, en route to skipping over empty lines

    :param value: value to check
    :type value: object

    :return: whether the value is filler
    :rtype: bool
    """
    if isinstance(value, str):
        return value.strip() == ''
    return value is None


def get_records(rows: Iterator[Any],
                fields: List[Any],
                primary_key_fields: List[str],
                choice_fields: Dict[str, Any]) -> List[Tuple[Any, Any]]:
    """
    Truncate/pad empty/missing records to expected row length, canonicalize
    cell content, and return resulting record list.

    :param upload_data: generator producing rows of content
    :type upload_data: generator
    :param fields: collection of fields specified in JSON schema
    :type fields: list or tuple
    :param primary_key_fields: list of field ids making up the PK
    :type primary_key_fields: list of strings
    :param choice_fields: {field_id: 'full'/True/False}
    :type choice_fields: dict

    :return: canonicalized records of specified upload data
    :rtype: tuple of dicts
    """
    records = []
    for n, row in rows:
        # trailing cells might be empty: trim row to fit
        while (row and
                (len(row) > len(fields)) and
                (row[-1] is None or row[-1] == '')):
            row.pop()
        while row and (len(row) < len(fields)):
            row.append(None)  # placeholder: canonicalize once only, below

        try:
            records.append(
                (n, dict((
                    f['datastore_id'],
                    canonicalize(
                        v,
                        f['datastore_type'],
                        f['datastore_id'] in primary_key_fields,
                        choice_fields.get(f['datastore_id'], False)))
                    for f, v in zip(fields, row))))
        except BadExcelData as e:
            raise BadExcelData('Row {0}:'.format(n) + ' ' + e.message)

    return records


# XXX remove this function once we upgrade to openpyxl 2.4
def unescape(value: str) -> str:
    """
    copy of unescape from openpyxl.utils.escape, openpyxl version 2.4.x
    """
    ESCAPED_REGEX = re.compile("_x([0-9A-Fa-f]{4})_")

    def _sub(match: Any):
        """
        Callback to unescape chars
        """
        return chr(int(match.group(1), 16))

    if "_x" in value:
        value = ESCAPED_REGEX.sub(_sub, value)

    return value
